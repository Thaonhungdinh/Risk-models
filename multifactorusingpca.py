"""
multifactor_risk_model_python.py

Full reproducible script that:
- Simulates 250 benchmark assets and a 50-asset portfolio (subset)
- Builds a 3-factor model (B exposures and factor covariance)
- Computes explained/unexplained systematic risk, specific risk, APT vol
- Computes tracking error decomposition (factor/explained, specific, market-systematic)
- Computes marginal contributions to tracking error (MC_to_TE)
- Exports an Excel workbook with Inputs, FactorCov, Outputs (formulas), MC_to_TE, Preferences
"""

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ---------------------------
# CONFIG
# ---------------------------
np.random.seed(0)
N_BENCH = 250       # benchmark universe
N_PORT = 50         # portfolio is subset of benchmark
N_FACTORS = 3
MARKET_VOL = 0.20   # annual market volatility (used for market-beta systematic)

# Output Excel filename
OUT_XLSX = "multifactor_risk_model_python.xlsx"

# ---------------------------
# SIMULATE ASSET UNIVERSE
# ---------------------------
assets = [f"Asset_{i+1}" for i in range(N_BENCH)]

# Market betas (used for "market-systematic" total systematic)
beta_market = np.random.uniform(0.8, 1.3, size=N_BENCH)

# Specific (idiosyncratic) volatilities (annual)
spec_vol = np.random.uniform(0.10, 0.30, size=N_BENCH)

# Factor exposures matrix B (N_BENCH x N_FACTORS)
B = np.random.normal(loc=0.0, scale=0.5, size=(N_BENCH, N_FACTORS))

# Build a positive-definite factor covariance matrix Sigma_f
A = np.random.normal(size=(N_FACTORS, N_FACTORS))
factor_cov = A @ A.T   # symmetric positive-definite

# ---------------------------
# BENCHMARK & PORTFOLIO WEIGHTS
# ---------------------------
# Benchmark: random weights normalized to sum 1
w_bench = np.random.random(N_BENCH)
w_bench = w_bench / w_bench.sum()

# Portfolio: equally-weighted subset of N_PORT assets (subset drawn from benchmark universe)
port_idx = np.random.choice(N_BENCH, size=N_PORT, replace=False)
w_port = np.zeros(N_BENCH)
w_port[port_idx] = 1.0 / N_PORT

# ---------------------------
# HELPER: compute exposures and variances
# ---------------------------
# Factor exposures of portfolio and benchmark: F_p = sum_i w_i * B_i
F_port = w_port @ B        # shape: (N_FACTORS,)
F_bench = w_bench @ B      # shape: (N_FACTORS,)

# Explained systematic variance (via factors): F' Σ_f F
explained_var_port = float(F_port @ factor_cov @ F_port.T)
explained_var_bench = float(F_bench @ factor_cov @ F_bench.T)

# Total systematic variance from market beta (market-systematic)
beta_port = float(w_port @ beta_market)
beta_bench = float(w_bench @ beta_market)
total_sys_var_port = (abs(beta_port) * MARKET_VOL)**2
total_sys_var_bench = (abs(beta_bench) * MARKET_VOL)**2

# Unexplained systematic variance = max(total_sys_var - explained_var, 0)
unexplained_var_port = max(total_sys_var_port - explained_var_port, 0.0)
unexplained_var_bench = max(total_sys_var_bench - explained_var_bench, 0.0)

# Specific variances: sum(w_i^2 * spec_i^2)
spec_var_port = float(np.sum((w_port * spec_vol)**2))
spec_var_bench = float(np.sum((w_bench * spec_vol)**2))

# Convert to vols
explained_vol_port = np.sqrt(max(explained_var_port, 0.0))
explained_vol_bench = np.sqrt(max(explained_var_bench, 0.0))
unexplained_vol_port = np.sqrt(unexplained_var_port)
unexplained_vol_bench = np.sqrt(unexplained_var_bench)
spec_vol_port_total = np.sqrt(spec_var_port)
spec_vol_bench_total = np.sqrt(spec_var_bench)

# APT total volatility (using market-systematic as "total systemic" per sheet)
apt_vol_port = np.sqrt(total_sys_var_port + spec_var_port)
apt_vol_bench = np.sqrt(total_sys_var_bench + spec_var_bench)

# ---------------------------
# TRACKING ERROR DECOMPOSITION
# ---------------------------
# Delta exposures
delta_beta = beta_port - beta_bench
delta_F = F_port - F_bench
delta_w = w_port - w_bench

# TE from factor exposures (explained systematic difference)
te_factor_var = float(delta_F @ factor_cov @ delta_F.T)
te_factor = np.sqrt(te_factor_var)

# TE specific from idiosyncratic differences: sqrt(sum((delta_w_i * spec_i)^2))
te_specific = np.sqrt(np.sum((delta_w * spec_vol)**2))

# TE market-systematic (difference in market beta times market vol)
te_market_systematic = abs(delta_beta) * MARKET_VOL

# TE totals:
te_total_market_decomp = np.sqrt(te_market_systematic**2 + te_specific**2)   # market-systematic + specific
te_total_factors_decomp = np.sqrt(te_factor**2 + te_specific**2)           # factor + specific
# Full TE including market-systematic + factor may double-count; choose decomposition based on desired approach.

# For completeness: TE from factors + market-systematic + specific (rarely used because market may be redundant)
# te_total_fullvar = np.sqrt(te_factor_var + np.sum((delta_w * spec_vol)**2) + (delta_beta * MARKET_VOL)**2)

# ---------------------------
# MARGINAL CONTRIBUTIONS TO TRACKING ERROR (MC_to_TE)
# ---------------------------
# We derive MC approximately via derivative of V = te_factor_var + sum((delta_w*spec)^2) + (delta_beta*mv)^2
# V = delta_F' Σ_f delta_F + sum((delta_w_i * spec_i)^2) + (delta_beta * MARKET_VOL)^2
# Note delta_F = sum_j delta_w_j * B_j  => derivative wrt w_i of delta_F is B_i (vector)
# dV/dw_i = 2 * (B_i^T Σ_f delta_F) + 2 * delta_w_i * spec_i^2 + 2 * beta_i * MARKET_VOL^2 * delta_beta
# Then MC_i = (1/(2 * TE_total)) * dV/dw_i  (because TE = sqrt(V) and d sqrt(V) = dV / (2 sqrt(V)))

# Compute V
V = te_factor_var + np.sum((delta_w * spec_vol)**2) + (delta_beta * MARKET_VOL)**2
sqrtV = np.sqrt(V) + 1e-12

# Precompute tmp = Σ_f delta_F (vector of factors)
tmp = factor_cov @ delta_F      # shape: (N_FACTORS,)

# dV/dw and MC_to_TE
dV_dw = np.zeros(N_BENCH)
for i in range(N_BENCH):
    # factor derivative: 2 * (B_i . (Σ_f delta_F))
    factor_der = 2.0 * (B[i, :] @ tmp)
    # specific derivative: 2 * delta_w_i * spec_i^2
    specific_der = 2.0 * delta_w[i] * (spec_vol[i]**2)
    # market derivative: 2 * beta_i * MARKET_VOL^2 * delta_beta
    market_der = 2.0 * beta_market[i] * (MARKET_VOL**2) * delta_beta
    dV_dw[i] = factor_der + specific_der + market_der

MC_to_TE = 0.5 * dV_dw / sqrtV

# For interpretation, create DataFrame of results
df = pd.DataFrame({
    "Asset": assets,
    "BenchWeight": w_bench,
    "PortWeight": w_port,
    "Delta_w": delta_w,
    "MarketBeta": beta_market,
    "SpecVol": spec_vol,
})
for k in range(N_FACTORS):
    df[f"Factor_{k+1}"] = B[:, k]
df["MC_to_TE"] = MC_to_TE
df["Abs_MC_to_TE"] = np.abs(MC_to_TE)

# Top 10 by absolute MC
preferences_top10 = df.sort_values("Abs_MC_to_TE", ascending=False).head(10).reset_index(drop=True)

# ---------------------------
# PRINT SUMMARY (console)
# ---------------------------
print("===== SUMMARY =====")
print(f"Portfolio # assets: {int(np.sum(w_port > 0))}  (subset indices: {np.where(w_port>0)[0][:10]} ...)")
print(f"Benchmark # assets: {N_BENCH}")
print(f"Portfolio beta: {beta_port:.6f}, Benchmark beta: {beta_bench:.6f}")
print(f"Explained sys var (port): {explained_var_port:.6e}, vol: {explained_vol_port:.6f}")
print(f"Unexplained sys var (port): {unexplained_var_port:.6e}, vol: {unexplained_vol_port:.6f}")
print(f"Specific var (port): {spec_var_port:.6e}, vol: {spec_vol_port_total:.6f}")
print(f"APT vol port (total): {apt_vol_port:.6f}")
print()
print("Tracking error decomposition (port vs benchmark):")
print(f"TE factor (explained): {te_factor:.6f}")
print(f"TE specific: {te_specific:.6f}")
print(f"TE market-systematic: {te_market_systematic:.6f}")
print(f"TE total (market decomposition): {te_total_market_decomp:.6f}")
print(f"TE total (factor decomposition): {te_total_factors_decomp:.6f}")
print()
print("Top 10 assets by absolute MC_to_TE:")
print(preferences_top10[["Asset", "Delta_w", "MC_to_TE"]].to_string(index=False))

# ---------------------------
# WRITE EXCEL WORKBOOK (Inputs, FactorCov, Outputs with formulas, MC_to_TE, Preferences)
# ---------------------------
wb = Workbook()
ws_inputs = wb.active
ws_inputs.title = "Inputs"

# Inputs header row
headers = ["Asset", "BenchWeight", "PortWeight", "MarketBeta", "SpecVol"] + [f"Factor_{i+1}" for i in range(N_FACTORS)]
ws_inputs.append(headers)

# Fill Inputs rows
for i in range(N_BENCH):
    row = [assets[i], float(w_bench[i]), float(w_port[i]), float(beta_market[i]), float(spec_vol[i])] + [float(B[i, k]) for k in range(N_FACTORS)]
    ws_inputs.append(row)

# Market vol cell for reference
ws_inputs["Z1"] = "Market_Vol"
ws_inputs["Z2"] = MARKET_VOL

# FactorCov sheet
ws_fcov = wb.create_sheet("FactorCov")
ws_fcov["A1"] = "FactorCov"
# Put header labels
for j in range(N_FACTORS):
    ws_fcov.cell(row=1, column=2+j, value=f"F{j+1}")
    ws_fcov.cell(row=2+j, column=1, value=f"F{j+1}")
# Fill covariance matrix into B2..
for i in range(N_FACTORS):
    for j in range(N_FACTORS):
        ws_fcov.cell(row=2+i, column=2+j, value=float(factor_cov[i, j]))

# Outputs sheet: write formulas referencing Inputs and FactorCov
ws_out = wb.create_sheet("Outputs")
ws_out["A1"] = "Metric"
ws_out["B1"] = "Portfolio"
ws_out["C1"] = "Benchmark"
ws_out["D1"] = "Notes"

metrics = [
    "Portfolio #Assets", "Benchmark #Assets",
    "Portfolio Beta", "Benchmark Beta",
    "Explained Sys Var (factors)", "Total Sys Var (market beta)",
    "Unexplained Sys Var", "Explained Sys Vol", "Unexplained Sys Vol",
    "Specific Var", "Specific Vol", "APT Vol (Total)",
    "TE Factor (sqrt)", "TE Specific (sqrt)", "TE Market-Systematic",
    "TE Total (market decomposition)", "TE Total (factor decomposition)"
]
for i, m in enumerate(metrics, start=2):
    ws_out[f"A{i}"] = m

# Basic counts (numbers)
ws_out["B2"] = int(np.sum(w_port > 0))
ws_out["C2"] = N_BENCH

# Portfolio Beta & Benchmark Beta formulas (Excel SUMPRODUCT)
# Input ranges: BenchWeight col B (2..N_BENCH+1), PortWeight col C (2..N_BENCH+1), MarketBeta col D (2..)
row_last = N_BENCH + 1
ws_out["B3"] = f"=SUMPRODUCT(Inputs!$C$2:$C${row_last}, Inputs!$D$2:$D${row_last})"  # portfolio beta
ws_out["C3"] = f"=SUMPRODUCT(Inputs!$B$2:$B${row_last}, Inputs!$D$2:$D${row_last})"  # benchmark beta

# Factor exposures: place Portfolio factor exposures in B5..B(5+N_FACTORS-1), same for Benchmark in C5..
start_row_fexp = 5
for k in range(N_FACTORS):
    col_letter = get_column_letter(6 + k)  # Factor_1 is column F (6)
    ws_out.cell(row=start_row_fexp + k, column=1).value = f"Port_Factor_{k+1}"
    ws_out.cell(row=start_row_fexp + k, column=2).value = f"=SUMPRODUCT(Inputs!$C$2:$C${row_last}, Inputs!${col_letter}$2:${col_letter}${row_last})"
    ws_out.cell(row=start_row_fexp + k, column=3).value = f"=SUMPRODUCT(Inputs!$B$2:$B${row_last}, Inputs!${col_letter}$2:${col_letter}${row_last})"

# Build addresses for MMULT formulas (explained variance)
pf_start = ws_out.cell(row=start_row_fexp, column=2).coordinate
pf_end = ws_out.cell(row=start_row_fexp + N_FACTORS - 1, column=2).coordinate
bf_start = ws_out.cell(row=start_row_fexp, column=3).coordinate
bf_end = ws_out.cell(row=start_row_fexp + N_FACTORS - 1, column=3).coordinate
fcov_start = ws_fcov.cell(row=2, column=2).coordinate
fcov_end = ws_fcov.cell(row=2 + N_FACTORS - 1, column=2 + N_FACTORS - 1).coordinate

# Place explained system var formulas at cells B(5+N_FACTORS) and C(...)
explained_row = start_row_fexp + N_FACTORS
ws_out.cell(row=explained_row, column=1).value = "Explained Sys Var (factors)"
ws_out.cell(row=explained_row, column=2).value = f"=MMULT(MMULT({pf_start}:{pf_end}, FactorCov!${fcov_start}:${fcov_end}), TRANSPOSE({pf_start}:{pf_end}))"
ws_out.cell(row=explained_row, column=3).value = f"=MMULT(MMULT({bf_start}:{bf_end}, FactorCov!${fcov_start}:${fcov_end}), TRANSPOSE({bf_start}:{bf_end}))"

# Total system var from market beta (BETA * MARKET_VOL)^2
ws_out.cell(row=explained_row+1, column=1).value = "Total Sys Var (market beta)"
ws_out.cell(row=explained_row+1, column=2).value = f"=({ws_out['B3'].coordinate} * Inputs!$Z$2)^2"
ws_out.cell(row=explained_row+1, column=3).value = f"=({ws_out['C3'].coordinate} * Inputs!$Z$2)^2"

# Unexplained sys var
ws_out.cell(row=explained_row+2, column=1).value = "Unexplained Sys Var"
ws_out.cell(row=explained_row+2, column=2).value = f"=MAX(B{explained_row+1} - B{explained_row}, 0)"
ws_out.cell(row=explained_row+2, column=3).value = f"=MAX(C{explained_row+1} - C{explained_row}, 0)"

# Explained / Unexplained vols
ws_out.cell(row=explained_row+3, column=1).value = "Explained Sys Vol"
ws_out.cell(row=explained_row+3, column=2).value = f"=SQRT(B{explained_row})"
ws_out.cell(row=explained_row+3, column=3).value = f"=SQRT(C{explained_row})"
ws_out.cell(row=explained_row+4, column=1).value = "Unexplained Sys Vol"
ws_out.cell(row=explained_row+4, column=2).value = f"=SQRT(B{explained_row+2})"
ws_out.cell(row=explained_row+4, column=3).value = f"=SQRT(C{explained_row+2})"

# Specific var and vol using SUMPRODUCT(w^2 * spec^2)
spec_var_row = explained_row + 5
ws_out.cell(row=spec_var_row, column=1).value = "Specific Var"
ws_out.cell(row=spec_var_row, column=2).value = f"=SUMPRODUCT(Inputs!$C$2:$C${row_last}*Inputs!$C$2:$C${row_last}, Inputs!$E$2:$E${row_last}*Inputs!$E$2:$E${row_last})"
ws_out.cell(row=spec_var_row, column=3).value = f"=SUMPRODUCT(Inputs!$B$2:$B${row_last}*Inputs!$B$2:$B${row_last}, Inputs!$E$2:$E${row_last}*Inputs!$E$2:$E${row_last})"
ws_out.cell(row=spec_var_row+1, column=1).value = "Specific Vol"
ws_out.cell(row=spec_var_row+1, column=2).value = f"=SQRT(B{spec_var_row})"
ws_out.cell(row=spec_var_row+1, column=3).value = f"=SQRT(C{spec_var_row})"

# APT Vol (Total) using Total Sys Var + Specific Var
apt_row = spec_var_row + 3
ws_out.cell(row=apt_row, column=1).value = "APT Vol (Total)"
ws_out.cell(row=apt_row, column=2).value = f"=SQRT(B{explained_row+1} + B{spec_var_row})"
ws_out.cell(row=apt_row, column=3).value = f"=SQRT(C{explained_row+1} + C{spec_var_row})"

# TE factor: compute delta factors rows placed under for readability
delta_f_start_row = apt_row + 2
for k in range(N_FACTORS):
    ws_out.cell(row=delta_f_start_row + k, column=1).value = f"Delta_F_{k+1}"
    # B col delta factor = Port_F_k - Bench_F_k
    ws_out.cell(row=delta_f_start_row + k, column=2).value = f"=B{start_row_fexp + k} - C{start_row_fexp + k}"

# TE Factor formula using MMULT on delta factor vector
df_start = ws_out.cell(row=delta_f_start_row, column=2).coordinate
df_end = ws_out.cell(row=delta_f_start_row + N_FACTORS - 1, column=2).coordinate
te_factor_row = delta_f_start_row + N_FACTORS
ws_out.cell(row=te_factor_row, column=1).value = "TE Factor (sqrt)"
ws_out.cell(row=te_factor_row, column=2).value = f"=SQRT( MMULT(MMULT({df_start}:{df_end}, FactorCov!${fcov_start}:${fcov_end}), TRANSPOSE({df_start}:{df_end})) )"

# TE Specific Excel formula analogous to Python: sqrt(SUM((delta_w * spec)^2))
te_specific_row = te_factor_row + 1
ws_out.cell(row=te_specific_row, column=1).value = "TE Specific (sqrt)"
ws_out.cell(row=te_specific_row, column=2).value = f"=SQRT( SUMPRODUCT( (Inputs!$C$2:$C${row_last}-Inputs!$B$2:$B${row_last})*(Inputs!$C$2:$C${row_last}-Inputs!$B$2:$B${row_last}), Inputs!$E$2:$E${row_last}*Inputs!$E$2:$E${row_last}) )"

# TE Market-systematic
te_market_row = te_specific_row + 1
ws_out.cell(row=te_market_row, column=1).value = "TE Market-Systematic"
ws_out.cell(row=te_market_row, column=2).value = f"=ABS(B3 - C3) * Inputs!$Z$2"

# TE totals
ws_out.cell(row=te_market_row+1, column=1).value = "TE Total (market decomposition)"
ws_out.cell(row=te_market_row+1, column=2).value = f"=SQRT(B{te_market_row}^2 + B{te_specific_row}^2)"
ws_out.cell(row=te_market_row+2, column=1).value = "TE Total (factor decomposition)"
ws_out.cell(row=te_market_row+2, column=2).value = f"=SQRT(B{te_factor_row}^2 + B{te_specific_row}^2)"

# EXPLANATORY NOTES
ws_out["D3"] = "Portfolio Beta = SUMPRODUCT(PortWeights, Market_Beta)"
ws_out["D" + str(explained_row)] = "Explained systematic variance = F_p * Sigma_f * F_p' where F_p = SUM(w_i * B_i)"
ws_out["D" + str(explained_row+2)] = "Unexplained systematic var = MAX(total_sys_var - explained_var, 0)"
ws_out["D" + str(spec_var_row)] = "Specific var = SUM(w_i^2 * spec_i^2)"
ws_out["D" + str(te_factor_row)] = "TE factor = sqrt( (F_p - F_b)' Σ_f (F_p - F_b) )"

# MC_to_TE sheet (numerical values written)
ws_mc = wb.create_sheet("MC_to_TE")
ws_mc.append(["Asset", "Delta_w", "MC_to_TE"])
for i in range(N_BENCH):
    ws_mc.append([assets[i], float(delta_w[i]), float(MC_to_TE[i])])

# Preferences sheet: top 10 MC by absolute value
ws_pref = wb.create_sheet("Preferences")
ws_pref.append(["Rank", "Asset", "Delta_w", "MC_to_TE"])
for i, row in preferences_top10.iterrows():
    ws_pref.append([i+1, row["Asset"], float(row["Delta_w"]), float(row["MC_to_TE"])])

# Save workbook
wb.save(OUT_XLSX)
print(f"\nExcel workbook saved to: {OUT_XLSX}")
